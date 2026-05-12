import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st
import io
import warnings

warnings.filterwarnings('ignore')

# =============================================================================
# 1. CONFIGURACIÓN GLOBAL
# =============================================================================
plt.rcParams.update({
    'figure.figsize': (10, 6),
    'font.size': 11,
    'axes.grid': True,
    'grid.alpha': 0.3,
    'savefig.dpi': 150,
    'savefig.bbox': 'tight',
    'figure.facecolor': 'white',
    'axes.facecolor': 'white'
})

RISK_FILLS = {
    'favorable': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'moderate':  PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'severe':    PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid'),
    'critical':  PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
}
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
SUB_FILL    = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
SUB_FONT    = Font(bold=True)

RISK_COLORS = {
    'favorable': '#C6EFCE',
    'moderate':  '#FFEB9C',
    'severe':    '#F4B084',
    'critical':  '#FFC7CE',
}
RISK_LABELS = {
    'favorable': '✅ FAVORABLE',
    'moderate':  '⚠️ MODERADO',
    'severe':    '🔶 SEVERO',
    'critical':  '🔴 CRÍTICO',
}

# =============================================================================
# 2. PARAM_INFO
# =============================================================================
PARAM_INFO = {
    'H':        {'label': 'Altura del banco (H)',              'unit': 'm',      'hint': 'Favorable >20 m | Típico 10-20 m | Desfavorable <10 m',            'default': 20.0,  'std_frac': 0.00, 'stochastic': False},
    'd':        {'label': 'Diámetro del barreno (d)',          'unit': 'mm',     'hint': 'Favorable 100-150 mm | Típico 76-100 mm | Desfavorable 50-76 mm',   'default': 76.2,  'std_frac': 0.02, 'stochastic': True},
    'rho_F':    {'label': 'Densidad explosivo fondo (ρ_F)',    'unit': 'kg/dm³', 'hint': 'Favorable 1.2-1.5 | Típico 1.0-1.2 | Desfavorable 0.8-1.0',        'default': 1.1,   'std_frac': 0.05, 'stochastic': True},
    'S_F':      {'label': 'Potencia relativa fondo (S_F)',     'unit': '-',      'hint': 'Favorable 1.1-1.3 | Típico 0.9-1.1 | Desfavorable 0.7-0.9',        'default': 1.0,   'std_frac': 0.05, 'stochastic': True},
    'rho_C':    {'label': 'Densidad explosivo columna (ρ_C)', 'unit': 'kg/dm³', 'hint': 'Favorable 1.1-1.3 | Típico 0.9-1.1 | Desfavorable 0.7-0.9',        'default': 0.8,   'std_frac': 0.08, 'stochastic': True},
    'E_V_ratio':{'label': 'Relación espaciamiento/burden (E/V)','unit': '-',     'hint': 'Favorable 0.8-1.0 | Típico 1.1-1.3 | Desfavorable 1.4-1.6',       'default': 1.25,  'std_frac': 0.05, 'stochastic': True},
    'f':        {'label': 'Factor de inclinación (f)',         'unit': '-',      'hint': 'Favorable 1.00 | Típico 0.85-0.95 | Desfavorable 0.70-0.85',       'default': 0.9,   'std_frac': 0.03, 'stochastic': True},
    'c':        {'label': 'Factor de roca (c)',                'unit': 'kg/m³',  'hint': 'Favorable 0.6-0.8 | Típico 0.4-0.6 | Desfavorable 0.2-0.4',       'default': 0.4,   'std_frac': 0.15, 'stochastic': True},
    'alpha':    {'label': 'Ángulo de inclinación (α)',         'unit': 'grados', 'hint': 'Favorable 0°-10° | Típico 10°-20° | Desfavorable 20°-30°',          'default': 18.43, 'std_frac': 0.10, 'stochastic': True},
    'n_sims':   {'label': 'Número de simulaciones Monte Carlo','unit': '-',      'hint': 'Estándar 10,000 | Alta precisión 50,000 | Baja 5,000',             'default': 10000, 'std_frac': 0.00, 'stochastic': False},
}

# =============================================================================
# 3. INTERPRETACIONES
# =============================================================================
def interpret_qe(qe_val):
    if qe_val < 0.25:
        return "Consumo muy bajo. Riesgo de fragmentación gruesa y pies del banco. Incrementar carga o reducir burden.", "critical"
    elif qe_val < 0.35:
        return "Consumo optimizado. Fragmentación balanceada, adecuada para carga y transporte estándar.", "favorable"
    elif qe_val < 0.50:
        return "Consumo moderado-alto. Posible sobre-fragmentación. Verificar confinamiento.", "moderate"
    else:
        return "Consumo crítico/elevado. Alta probabilidad de vibraciones excesivas y costos descontrolados.", "severe"

def interpret_rp(rp_val):
    if rp_val > 12.0:
        return "Rendimiento excelente. Alta eficiencia en perforación y bajo costo operativo.", "favorable"
    elif rp_val > 9.0:
        return "Rendimiento aceptable. Operación dentro de rangos industriales típicos.", "moderate"
    elif rp_val > 6.0:
        return "Rendimiento bajo. Posible ineficiencia en diseño o roca adversa. Revisar malla.", "severe"
    else:
        return "Rendimiento crítico. Diseño ineficiente o parámetros inviables. Ajustar geometría.", "critical"

def interpret_hc(hc_val):
    if hc_val <= 0.5:
        return "Columna crítica/nula. Confinamiento insuficiente. Riesgo de disparo prematuro.", "critical"
    elif hc_val <= 3.0:
        return "Columna reducida. Ajustar retacado o reducir burden.", "severe"
    else:
        return "Longitud de columna adecuada. Distribución de energía favorable.", "favorable"

# =============================================================================
# 4. CÁLCULOS
# =============================================================================
def calc_V(d, rho_F, S_F, c, f, E_V_ratio):
    return (d / 33.0) * np.sqrt((rho_F * S_F) / (c * f * E_V_ratio))

def calc_Vp(V): return 0.9 * V
def calc_E(Vp, E_V_ratio): return E_V_ratio * Vp
def calc_U(Vp): return 0.3 * Vp
def calc_L(U, H, alpha_deg): return U + H / np.cos(np.radians(alpha_deg))
def calc_hf(Vp): return 1.3 * Vp
def calc_hr(Vp): return Vp
def calc_hc(L, hf, hr): return max(L - hf - hr, 0.0)

def calc_Q_and_outputs(p, L, hf, hc, Vp, E):
    d_m = p['d'] / 1000.0
    A = np.pi * (d_m ** 2) / 4.0
    q_f = A * hf * p['rho_F'] * 1000.0
    q_c = A * hc * p['rho_C'] * 1000.0
    Q = q_f + q_c
    vol = Vp * E * p['H']
    Qe = Q / vol if vol != 0 else 0.0
    Rp = vol / L if L != 0 else 0.0
    return Q, q_f, q_c, Qe, Rp

def run_deterministic(p):
    V  = calc_V(p['d'], p['rho_F'], p['S_F'], p['c'], p['f'], p['E_V_ratio'])
    Vp = calc_Vp(V)
    E  = calc_E(Vp, p['E_V_ratio'])
    U  = calc_U(Vp)
    L  = calc_L(U, p['H'], p['alpha'])
    hf = calc_hf(Vp)
    hr = calc_hr(Vp)
    hc = calc_hc(L, hf, hr)
    Q, q_f, q_c, Qe, Rp = calc_Q_and_outputs(p, L, hf, hc, Vp, E)
    return {'V': V, 'Vp': Vp, 'E': E, 'U': U, 'L': L, 'hf': hf, 'hr': hr, 'hc': hc,
            'Q': Q, 'q_f': q_f, 'q_c': q_c, 'Qe': Qe, 'Rp': Rp}

def run_monte_carlo(p):
    n = int(p['n_sims'])
    samples = {}
    for k, info in PARAM_INFO.items():
        if info['stochastic']:
            mean = p[k]; std = p[k] * info['std_frac']
            a = -mean / std if std > 0 else -np.inf
            samples[k] = stats.truncnorm.rvs(a, np.inf, loc=mean, scale=std, size=n)
        else:
            samples[k] = np.full(n, p[k])

    V  = (samples['d'] / 33.0) * np.sqrt((samples['rho_F'] * samples['S_F']) /
                                           (samples['c'] * samples['f'] * samples['E_V_ratio']))
    Vp = 0.9 * V
    E  = samples['E_V_ratio'] * Vp
    U  = 0.3 * Vp
    L  = U + samples['H'] / np.cos(np.radians(samples['alpha']))
    hf = 1.3 * Vp
    hr = Vp
    hc = np.maximum(L - hf - hr, 0.0)

    d_m = samples['d'] / 1000.0
    A   = np.pi * d_m**2 / 4.0
    q_f = A * hf * samples['rho_F'] * 1000.0
    q_c = A * hc * samples['rho_C'] * 1000.0
    Q   = q_f + q_c
    vol = Vp * E * samples['H']
    Qe  = np.divide(Q,   vol, out=np.zeros_like(Q),   where=vol != 0)
    Rp  = np.divide(vol, L,   out=np.zeros_like(vol), where=L   != 0)
    Qe  = np.where(np.isfinite(Qe), Qe, 0.0)
    Rp  = np.where(np.isfinite(Rp), Rp, 0.0)

    mc_res = {'V': V, 'Vp': Vp, 'E': E, 'U': U, 'L': L,
              'hf': hf, 'hr': hr, 'hc': hc, 'Q': Q,
              'q_f': q_f, 'q_c': q_c, 'Qe': Qe, 'Rp': Rp}
    return samples, mc_res

def compute_mc_stats(mc_res):
    mc_stats = {}
    for var in ['V', 'Vp', 'E', 'U', 'L', 'hf', 'hr', 'hc', 'Q', 'Qe', 'Rp']:
        arr = mc_res[var]
        mc_stats[f'{var}_mean'] = np.mean(arr)
        mc_stats[f'{var}_std']  = np.std(arr)
        mc_stats[f'{var}_P5']   = np.percentile(arr, 5)
        mc_stats[f'{var}_P50']  = np.percentile(arr, 50)
        mc_stats[f'{var}_P90']  = np.percentile(arr, 90)
        mc_stats[f'{var}_P95']  = np.percentile(arr, 95)
    return mc_stats

# =============================================================================
# 5. GRÁFICAS → bytes (en memoria)
# =============================================================================
def fig_to_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    return buf.read()

def make_plots(det, mc_res, mc_samples, p):
    plots = {}

    # Fig 1: Histograma + CDF Qe
    fig, ax1 = plt.subplots()
    ax2 = ax1.twinx()
    ax1.hist(mc_res['Qe'], bins=50, color='skyblue', edgecolor='black', density=True, alpha=0.6, label='PDF')
    sorted_qe = np.sort(mc_res['Qe'])
    cdf = np.arange(1, len(sorted_qe) + 1) / len(sorted_qe)
    ax2.plot(sorted_qe, cdf, 'r-', linewidth=2, label='CDF')
    ax1.axvline(det['Qe'], color='green', linestyle='--', label=f'Det: {det["Qe"]:.3f}')
    ax1.set_xlabel('Consumo Específico Qe (kg/m³)')
    ax1.set_ylabel('Densidad')
    ax2.set_ylabel('Probabilidad Acumulada')
    ax1.legend(loc='upper left', fontsize=8)
    ax2.legend(loc='upper right', fontsize=8)
    plt.title('Fig. 1 – Distribución y CDF del Consumo Específico (Monte Carlo)')
    plots['fig1'] = fig_to_bytes(fig)
    plt.close()

    # Fig 2: Histograma + CDF Rp
    fig, ax1 = plt.subplots()
    ax2 = ax1.twinx()
    ax1.hist(mc_res['Rp'], bins=50, color='lightgreen', edgecolor='black', density=True, alpha=0.6, label='PDF')
    sorted_rp = np.sort(mc_res['Rp'])
    cdf_rp = np.arange(1, len(sorted_rp) + 1) / len(sorted_rp)
    ax2.plot(sorted_rp, cdf_rp, 'purple', linewidth=2, label='CDF')
    ax1.axvline(det['Rp'], color='green', linestyle='--', label=f'Det: {det["Rp"]:.2f}')
    ax1.set_xlabel('Rendimiento Perforación Rp (m³/m)')
    ax1.set_ylabel('Densidad')
    ax2.set_ylabel('Probabilidad Acumulada')
    ax1.legend(loc='upper left', fontsize=8)
    ax2.legend(loc='upper right', fontsize=8)
    plt.title('Fig. 2 – Distribución y CDF del Rendimiento de Perforación (Monte Carlo)')
    plots['fig2'] = fig_to_bytes(fig)
    plt.close()

    # Fig 3: Tornado de Sensibilidad
    inputs_stoch = ['c', 'd', 'rho_F', 'rho_C', 'E_V_ratio', 'f', 'S_F']
    corrs = [np.corrcoef(mc_samples[v], mc_res['Qe'])[0, 1] for v in inputs_stoch]
    fig, ax = plt.subplots(figsize=(10, 5))
    y_pos  = np.arange(len(inputs_stoch))
    colors = ['red' if c < 0 else 'steelblue' for c in corrs]
    ax.barh(y_pos, corrs, color=colors, edgecolor='black')
    ax.set_yticks(y_pos)
    ax.set_yticklabels(inputs_stoch)
    ax.set_xlabel('Coeficiente de Correlación de Pearson con Qe')
    ax.axvline(0, color='black', linewidth=0.8)
    plt.title('Fig. 3 – Diagrama de Tornado de Sensibilidad (Correlación vs Qe)')
    plt.tight_layout()
    plots['fig3'] = fig_to_bytes(fig)
    plt.close()

    # Fig 4: Scatter c vs Qe
    fig, ax = plt.subplots()
    ax.scatter(mc_samples['c'], mc_res['Qe'], alpha=0.3, s=10, color='gray', label='Simulaciones')
    c_vals  = np.linspace(0.1, 1.0, 100)
    V_curve = (p['d'] / 33.0) * np.sqrt((p['rho_F'] * p['S_F']) / (c_vals * p['f'] * p['E_V_ratio']))
    Vp_c    = 0.9 * V_curve
    E_c     = p['E_V_ratio'] * Vp_c
    vol_c   = Vp_c * E_c * p['H']
    Qe_c    = det['Q'] / vol_c
    ax.plot(c_vals, Qe_c, 'r-', linewidth=2, label='Tendencia Determinista')
    ax.set_xlabel('Factor de Roca c (kg/m³)')
    ax.set_ylabel('Consumo Específico Qe (kg/m³)')
    plt.title('Fig. 4 – Relación Factor de Roca vs Consumo Específico')
    plt.legend()
    plots['fig4'] = fig_to_bytes(fig)
    plt.close()

    return plots

# =============================================================================
# 6. EXPORT EXCEL → bytes
# =============================================================================
def export_excel_bytes(p, det, mc_stats, mc_samples, mc_results):
    wb = Workbook()
    ws1 = wb.create_sheet("1_Datos_Entrada")
    ws2 = wb.create_sheet("2_Interpretacion_Resultados")
    ws3 = wb.create_sheet("3_Diseno_Voladura")
    ws4 = wb.create_sheet("4_MonteCarlo_Estadisticas")
    ws5 = wb.create_sheet("5_MonteCarlo_Muestra")

    def apply_header(ws, row, cols):
        for i, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border    = Border(bottom=Side(style='thin'))

    def apply_row(ws, row, data):
        for i, val in enumerate(data, 1):
            ws.cell(row=row, column=i, value=val)

    # Hoja 1
    apply_header(ws1, 1, ['Parámetro', 'Valor', 'Unidades', 'Valores Típicos', 'CoV MC (%)'])
    for r, k in enumerate(PARAM_INFO, 2):
        info = PARAM_INFO[k]
        apply_row(ws1, r, [info['label'], p[k], info['unit'], info['hint'], info['std_frac'] * 100])
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['D'].width = 60

    # Hoja 2
    apply_header(ws2, 1, ['Resultado', 'Valor Determinista', 'Interpretación', 'Nivel de Riesgo'])
    interp_data = [
        ('Piedra Práctica Vp',        f"{det['Vp']:.2f} m",      "Distancia crítica ajustada para fragmentación óptima.",                 "moderate"),
        ('Espaciamiento E',           f"{det['E']:.2f} m",       "Distancia entre taladros para cobertura uniforme del frente.",           "moderate"),
        ('Consumo Específico Qe',     f"{det['Qe']:.3f} kg/m³",  *interpret_qe(det['Qe'])),
        ('Rendimiento Perforación Rp',f"{det['Rp']:.2f} m³/m",   *interpret_rp(det['Rp'])),
        ('Longitud Columna hc',       f"{det['hc']:.2f} m",      *interpret_hc(det['hc'])),
    ]
    for r, (res, val, interp, risk) in enumerate(interp_data, 2):
        apply_row(ws2, r, [res, val, interp, risk.upper()])
        ws2.cell(row=r, column=2).fill = RISK_FILLS[risk]
        ws2.cell(row=r, column=4).fill = RISK_FILLS[risk]
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['C'].width = 65

    # Hoja 3
    apply_header(ws3, 1, ['Paso', 'Parámetro', 'Símbolo', 'Fórmula', 'Valor'])
    steps = [
        ('1','Piedra Máxima Teórica','V',  'V=(d/33)*sqrt((ρF·SF)/(c·f·(E/V)))', det['V']),
        ('2','Piedra Práctica',      'Vp', 'Vp=0.9·V',                            det['Vp']),
        ('3','Espaciamiento',        'E',  'E=(E/V)·Vp',                          det['E']),
        ('4','Sobreperforación',     'U',  'U=0.3·Vp',                            det['U']),
        ('5','Long. Perforación',    'L',  'L=U+H/cos(α)',                         det['L']),
        ('6','Carga Fondo',          'hf', 'hf=1.3·Vp',                           det['hf']),
        ('7','Retacado',             'hr', 'hr=Vp',                                det['hr']),
        ('8','Carga Columna',        'hc', 'hc=L−hf−hr',                          det['hc']),
        ('9','Carga Total',          'Q',  'Q=qf+qc',                             det['Q']),
        ('10','Consumo Específico',  'Qe', 'Qe=Q/(Vp·E·H)',                       det['Qe']),
        ('11','Rendimiento',         'Rp', 'Rp=(Vp·E·H)/L',                       det['Rp']),
    ]
    for i, s in enumerate(steps, 2):
        apply_row(ws3, i, s)

    # Hoja 4
    apply_header(ws4, 1, ['Variable', 'Media', 'Desv. Est.', 'P5', 'P50', 'P90', 'P95'])
    for idx, var in enumerate(['V','Vp','E','U','L','hf','hr','hc','Q','Qe','Rp'], 2):
        apply_row(ws4, idx, [var, mc_stats[f'{var}_mean'], mc_stats[f'{var}_std'],
                             mc_stats[f'{var}_P5'], mc_stats[f'{var}_P50'],
                             mc_stats[f'{var}_P90'], mc_stats[f'{var}_P95']])
    apply_row(ws4, 15, ['Correlación Pearson vs Qe'])
    ws4.cell(row=15, column=1).fill = SUB_FILL
    ws4.cell(row=15, column=1).font = SUB_FONT
    apply_header(ws4, 16, ['Variable', 'Correlación'])
    for i, var in enumerate(['c','d','rho_F','rho_C','E_V_ratio','f','S_F','alpha'], 17):
        corr = np.corrcoef(mc_samples[var], mc_results['Qe'])[0, 1]
        apply_row(ws4, i, [var, corr])

    # Hoja 5
    headers = list(mc_samples.keys()) + list(mc_results.keys())
    apply_header(ws5, 1, headers)
    for i in range(min(500, len(mc_results['Qe']))):
        row_data = [mc_samples[k][i] for k in mc_samples] + [mc_results[k][i] for k in mc_results]
        apply_row(ws5, i + 2, row_data)

    # Eliminar hoja vacía por defecto
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# =============================================================================
# 7. STREAMLIT UI
# =============================================================================
def main():
    st.set_page_config(
        page_title="Diseño Malla de Voladura – Langefors-Kihlström",
        page_icon="💥",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # ── Estilos ──────────────────────────────────────────────────────────────
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700&family=IBM+Plex+Mono:wght@400;600&display=swap');

    html, body, [class*="css"] { font-family: 'Barlow Condensed', sans-serif; }
    code, .stCode { font-family: 'IBM Plex Mono', monospace; }

    .block-container { padding-top: 1.5rem; }

    .metric-card {
        background: #0d1b2a;
        border: 1px solid #1F4E79;
        border-radius: 8px;
        padding: 14px 18px;
        text-align: center;
        color: #e0e0e0;
        margin-bottom: 8px;
    }
    .metric-card .val {
        font-size: 2rem;
        font-weight: 700;
        color: #4FC3F7;
        display: block;
        line-height: 1.1;
    }
    .metric-card .lbl { font-size: 0.85rem; color: #90CAF9; letter-spacing: 0.05em; }

    .risk-badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 4px;
        font-weight: 700;
        font-size: 0.85rem;
    }
    .interp-box {
        background: #0d1b2a;
        border-left: 4px solid #1F4E79;
        border-radius: 0 6px 6px 0;
        padding: 10px 14px;
        margin: 4px 0 12px;
        color: #cfd8dc;
        font-size: 0.92rem;
    }

    .st-expander { border: 1px solid #1F4E79 !important; border-radius: 6px; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] { 
        background: #0d1b2a; 
        border: 1px solid #1F4E79; 
        border-radius: 6px 6px 0 0;
        color: #90CAF9;
        font-weight: 600;
        font-size: 0.9rem;
    }
    .stTabs [aria-selected="true"] { background: #1F4E79 !important; color: #fff !important; }

    h1, h2, h3 { font-family: 'Barlow Condensed', sans-serif; }
    </style>
    """, unsafe_allow_html=True)

    # ── Encabezado ───────────────────────────────────────────────────────────
    st.markdown("""
    <div style="background:linear-gradient(135deg,#0d1b2a 0%,#1F4E79 100%);
                padding:24px 32px;border-radius:10px;margin-bottom:24px;">
      <h1 style="color:#4FC3F7;margin:0;font-size:2.2rem;letter-spacing:0.04em;">
        💥 DISEÑO DE MALLA DE PERFORACIÓN Y VOLADURA
      </h1>
      <p style="color:#90CAF9;margin:4px 0 0;font-size:1.05rem;">
        Método de <strong>Langefors-Kihlström</strong> · Análisis Determinista + Monte Carlo
      </p>
    </div>
    """, unsafe_allow_html=True)

    # ── Sidebar: Parámetros de entrada ───────────────────────────────────────
    with st.sidebar:
        st.markdown("## ⚙️ Parámetros de Entrada")

        sections = [
            ('🏔️ Geometría del Banco',          ['H', 'd']),
            ('💣 Propiedades del Explosivo',      ['rho_F', 'S_F', 'rho_C']),
            ('📐 Parámetros Diseño y Roca',       ['E_V_ratio', 'f', 'c', 'alpha']),
            ('🎲 Configuración Monte Carlo',      ['n_sims']),
        ]

        p = {}
        for sec_name, keys in sections:
            with st.expander(sec_name, expanded=True):
                for k in keys:
                    info = PARAM_INFO[k]
                    st.caption(f"💡 {info['hint']}")
                    if k == 'n_sims':
                        p[k] = st.number_input(
                            f"{info['label']} [{info['unit']}]",
                            min_value=1000, max_value=100000,
                            value=int(info['default']), step=1000, key=k
                        )
                    else:
                        p[k] = st.number_input(
                            f"{info['label']} [{info['unit']}]",
                            value=float(info['default']),
                            format="%.4f", key=k
                        )

        run_btn = st.button("▶ CALCULAR", type="primary", use_container_width=True)

    # ── Cuerpo principal ─────────────────────────────────────────────────────
    if not run_btn:
        st.info("👈 Ajusta los parámetros en el panel izquierdo y presiona **CALCULAR**.")
        st.markdown("""
        ### ¿Qué hace esta aplicación?
        - Calcula la geometría óptima de malla de perforación según Langefors-Kihlström.
        - Realiza un análisis probabilístico mediante **Simulación de Monte Carlo**.
        - Genera gráficas de distribución, CDF y análisis de sensibilidad (Tornado).
        - Exporta todos los resultados en un archivo **Excel** descargable.
        """)
        return

    # ── Cálculos ─────────────────────────────────────────────────────────────
    with st.spinner("Ejecutando cálculos..."):
        det = run_deterministic(p)
        samples, mc_res = run_monte_carlo(p)
        mc_stats = compute_mc_stats(mc_res)
        plots    = make_plots(det, mc_res, samples, p)
        excel_b  = export_excel_bytes(p, det, mc_stats, samples, mc_res)

    st.success(f"✅ Cálculo completado · {int(p['n_sims']):,} simulaciones Monte Carlo ejecutadas.")

    # ── Tabs ─────────────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Resultados Deterministas",
        "📈 Análisis Monte Carlo",
        "🌪️ Sensibilidad",
        "📥 Exportar",
    ])

    # ── Tab 1: Resultados Deterministas ──────────────────────────────────────
    with tab1:
        st.markdown("### 🔩 Geometría del Barreno")
        c1, c2, c3, c4 = st.columns(4)
        metrics = [
            (c1, "Piedra Práctica Vp", f"{det['Vp']:.2f}", "m"),
            (c2, "Espaciamiento E",    f"{det['E']:.2f}",  "m"),
            (c3, "Long. Perforación L",f"{det['L']:.2f}",  "m"),
            (c4, "Sobreperforación U", f"{det['U']:.2f}",  "m"),
        ]
        for col, label, val, unit in metrics:
            col.markdown(f"""
            <div class="metric-card">
              <span class="val">{val}</span>
              <span class="lbl">{label} ({unit})</span>
            </div>""", unsafe_allow_html=True)

        st.markdown("### 💣 Carga Explosiva")
        c1, c2, c3, c4 = st.columns(4)
        metrics2 = [
            (c1, "Carga Total Q",      f"{det['Q']:.1f}",   "kg"),
            (c2, "Carga Fondo qf",     f"{det['q_f']:.1f}", "kg"),
            (c3, "Carga Columna qc",   f"{det['q_c']:.1f}", "kg"),
            (c4, "Long. Columna hc",   f"{det['hc']:.2f}",  "m"),
        ]
        for col, label, val, unit in metrics2:
            col.markdown(f"""
            <div class="metric-card">
              <span class="val">{val}</span>
              <span class="lbl">{label} ({unit})</span>
            </div>""", unsafe_allow_html=True)

        st.markdown("### 🎯 Indicadores de Rendimiento")
        for result_data in [
            ('Consumo Específico Qe', f"{det['Qe']:.3f} kg/m³", *interpret_qe(det['Qe'])),
            ('Rendimiento Perforación Rp', f"{det['Rp']:.2f} m³/m", *interpret_rp(det['Rp'])),
            ('Longitud Columna hc', f"{det['hc']:.2f} m", *interpret_hc(det['hc'])),
        ]:
            name, val, interp, risk = result_data
            badge_color = RISK_COLORS[risk]
            badge_label = RISK_LABELS[risk]
            st.markdown(f"""
            **{name}** — <span style="font-size:1.1rem;font-weight:700;color:#4FC3F7;">{val}</span>
            &nbsp;<span class="risk-badge" style="background:{badge_color};color:#000;">{badge_label}</span>
            <div class="interp-box">{interp}</div>
            """, unsafe_allow_html=True)

        st.markdown("### 📋 Secuencia de Cálculo (Langefors-Kihlström)")
        steps_data = {
            'Paso':      ['1','2','3','4','5','6','7','8','9','10','11'],
            'Parámetro': ['Piedra Máxima V','Piedra Práctica Vp','Espaciamiento E',
                          'Sobreperf. U','Long. Perf. L','Carga Fondo hf',
                          'Retacado hr','Carga Columna hc','Carga Total Q',
                          'Consumo Específico Qe','Rendimiento Rp'],
            'Fórmula':   ['(d/33)·√(ρF·SF/c·f·E/V)','0.9·V','(E/V)·Vp',
                          '0.3·Vp','U+H/cos(α)','1.3·Vp',
                          'Vp','L−hf−hr','qf+qc',
                          'Q/(Vp·E·H)','(Vp·E·H)/L'],
            'Valor':     [f"{det['V']:.3f} m",f"{det['Vp']:.3f} m",f"{det['E']:.3f} m",
                          f"{det['U']:.3f} m",f"{det['L']:.3f} m",f"{det['hf']:.3f} m",
                          f"{det['hr']:.3f} m",f"{det['hc']:.3f} m",f"{det['Q']:.2f} kg",
                          f"{det['Qe']:.4f} kg/m³",f"{det['Rp']:.2f} m³/m"],
        }
        st.dataframe(steps_data, use_container_width=True, hide_index=True)

    # ── Tab 2: Monte Carlo ───────────────────────────────────────────────────
    with tab2:
        st.markdown("### 📊 Estadísticas Monte Carlo")
        mc_table = {
            'Variable': ['V','Vp','E','U','L','hf','hr','hc','Q','Qe','Rp'],
        }
        for stat in ['mean','std','P5','P50','P90','P95']:
            mc_table[stat] = [round(mc_stats[f'{v}_{stat}'], 4) for v in mc_table['Variable']]
        st.dataframe(mc_table, use_container_width=True, hide_index=True)

        st.markdown("### 📈 Gráficas de Distribución")
        col1, col2 = st.columns(2)
        with col1:
            st.image(plots['fig1'], caption="Fig. 1 – PDF y CDF del Consumo Específico Qe", use_container_width=True)
        with col2:
            st.image(plots['fig2'], caption="Fig. 2 – PDF y CDF del Rendimiento de Perforación Rp", use_container_width=True)

    # ── Tab 3: Sensibilidad ──────────────────────────────────────────────────
    with tab3:
        st.markdown("### 🌪️ Análisis de Sensibilidad")
        col1, col2 = st.columns(2)
        with col1:
            st.image(plots['fig3'], caption="Fig. 3 – Diagrama de Tornado (Correlación vs Qe)", use_container_width=True)
        with col2:
            st.image(plots['fig4'], caption="Fig. 4 – Factor de Roca c vs Consumo Específico Qe", use_container_width=True)

        st.markdown("### 📊 Tabla de Correlaciones (Pearson con Qe)")
        input_vars = ['c','d','rho_F','rho_C','E_V_ratio','f','S_F','alpha']
        corr_table = {
            'Variable': input_vars,
            'Correlación con Qe': [round(np.corrcoef(samples[v], mc_res['Qe'])[0,1], 4) for v in input_vars],
        }
        st.dataframe(corr_table, use_container_width=True, hide_index=True)

    # ── Tab 4: Exportar ──────────────────────────────────────────────────────
    with tab4:
        st.markdown("### 📥 Descargar Resultados")
        st.info("El archivo Excel contiene 5 hojas: Datos de Entrada, Interpretación, Secuencia de Cálculo, Estadísticas Monte Carlo y Muestra de Simulaciones.")
        st.download_button(
            label="⬇️ Descargar Excel – Resultados_Voladura_Langefors.xlsx",
            data=excel_b,
            file_name="Resultados_Voladura_Langefors.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


if __name__ == "__main__":
    main()
